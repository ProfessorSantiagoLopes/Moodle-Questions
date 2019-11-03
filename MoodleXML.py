#!/usr/bin/python
# encoding: utf-8
from __future__ import print_function
from openpyxl import load_workbook
import sys

# Nome do arquivo Excel
wb = load_workbook(filename = sys.argv[1])
ws = wb.active

# Gravar o arquivo .xml
f= open(sys.argv[1]+".xml","w+")

# Cabeçalho do quiz
quizHeader = """
<?xml version="1.0" encoding="UTF-8"?>
<quiz>"""

# Cabeçalho da questão
questionHeader = """
    <question type="multichoice">
        <name>
            <text>{0}-{1}-Q"""
# Questão
question = """{0}</text>
        </name>
        <questiontext format="html">
            <text><![CDATA[<p>{1}</p>]]></text>
        </questiontext>
        <generalfeedback format="html">
            <text></text>
        </generalfeedback>
        <defaultgrade>1.0000000</defaultgrade>
        <penalty>0.3333333</penalty>
        <hidden>0</hidden>
        <single>true</single>
        <shuffleanswers>true</shuffleanswers>
        <answernumbering>abc</answernumbering>
        <correctfeedback format="html">
            <text></text>
        </correctfeedback>
        <partiallycorrectfeedback format="html">
            <text></text>
        </partiallycorrectfeedback>
        <incorrectfeedback format="html">
            <text></text>
        </incorrectfeedback>
        <answer fraction="100" format="html">
            <text><![CDATA[<p><span>{2}</span></p>]]></text>
            <feedback format="html">
            <text></text>
            </feedback>
        </answer>
        <answer fraction="0" format="html">
            <text><![CDATA[<p><span>{3}</span></p>]]></text>
            <feedback format="html">
            <text></text>
            </feedback>
        </answer>
        <answer fraction="0" format="html">
            <text><![CDATA[<p><span>{4}</span></p>]]></text>
            <feedback format="html">
            <text></text>
            </feedback>
        </answer>
        <answer fraction="0" format="html">
            <text><![CDATA[<p><span>{5}</span></p>]]></text>
            <feedback format="html">
            <text></text>
            </feedback>
        </answer>
    </question>
"""
# Rodapé do questionário
quizFooter ="""</quiz>"""

B1 = str(ws["B1"].value)
B2 = str(ws["B2"].value)

if (B2) :
    print(quizHeader, end = '')
if (B2) :
    f.write(quizHeader)

A5 = str(ws["A5"].value)
B5 = ws["B5"].value
C5 = ws["C5"].value
D5 = ws["D5"].value
E5 = ws["E5"].value
F5 = ws["F5"].value
G5 = ws["G5"].value
if (B5) :
    print(questionHeader.format(B1, B2), end = '')
if (B5) :
    f.write(questionHeader.format(B1, B2))
if (B5) :
    print(question.format(A5.encode('utf8'), B5.encode('utf8'), C5.encode('utf8'), D5.encode('utf8'), E5.encode('utf8'), F5.encode('utf8'), G5.encode('utf8')), end = '')
if (B5) :
    f.write(question.format(A5.encode('utf8'), B5.encode('utf8'), C5.encode('utf8'), D5.encode('utf8'), E5.encode('utf8'), F5.encode('utf8'), G5.encode('utf8')))

A6 = str(ws["A6"].value)
B6 = ws["B6"].value
C6 = ws["C6"].value
D6 = ws["D6"].value
E6 = ws["E6"].value
F6 = ws["F6"].value
G6 = ws["G6"].value
if (B6) :
    print(questionHeader.format(B1, B2), end = '')
if (B6) :
    f.write(questionHeader.format(B1, B2))
if (B6) :
    print(question.format(A6.encode('utf8'), B6.encode('utf8'), C6.encode('utf8'), D6.encode('utf8'), E6.encode('utf8'), F6.encode('utf8'), G6.encode('utf8')), end = '')
if (B6) :
    f.write(question.format(A6.encode('utf8'), B6.encode('utf8'), C6.encode('utf8'), D6.encode('utf8'), E6.encode('utf8'), F6.encode('utf8'), G6.encode('utf8')))

A7 = str(ws["A7"].value)
B7 = ws["B7"].value
C7 = ws["C7"].value
D7 = ws["D7"].value
E7 = ws["E7"].value
F7 = ws["F7"].value
G7 = ws["G7"].value
if (B7) :
    print(questionHeader.format(B1, B2), end = '')
if (B7) :
    f.write(questionHeader.format(B1, B2))
if (B7) :
    print(question.format(A7.encode('utf8'), B7.encode('utf8'), C7.encode('utf8'), D7.encode('utf8'), E7.encode('utf8'), F7.encode('utf8'), G7.encode('utf8')), end = '')
if (B7) :
    f.write(question.format(A7.encode('utf8'), B7.encode('utf8'), C7.encode('utf8'), D7.encode('utf8'), E7.encode('utf8'), F7.encode('utf8'), G7.encode('utf8')))

A8 = str(ws["A8"].value)
B8 = ws["B8"].value
C8 = ws["C8"].value
D8 = ws["D8"].value
E8 = ws["E8"].value
F8 = ws["F8"].value
G8 = ws["G8"].value
if (B8) :
    print(questionHeader.format(B1, B2), end = '')
if (B8) :
    f.write(questionHeader.format(B1, B2))
if (B8) :
    print(question.format(A8.encode('utf8'), B8.encode('utf8'), C8.encode('utf8'), D8.encode('utf8'), E8.encode('utf8'), F8.encode('utf8'), G8.encode('utf8')), end = '')
if (B8) :
    f.write(question.format(A8.encode('utf8'), B8.encode('utf8'), C8.encode('utf8'), D8.encode('utf8'), E8.encode('utf8'), F8.encode('utf8'), G8.encode('utf8')))

print(quizFooter)
